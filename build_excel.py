#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Genera un Excel con las vacantes scrapeadas (OCC + Computrabajo) de esta sesion,
deduplicadas, con enlace y una PUNTUACION DE COINCIDENCIA (0-100) calculada contra
el CV de Claudia Elena Pedrozo Machorro:
  - Perfil: SEO, Marketing de Contenidos, Comunicacion Interna/Corporativa,
    Community/Social Media, Email Marketing, Google Analytics/Looker, WordPress,
    Photoshop/Illustrator/Premiere/Canva. Bilingue (esp/ing avanzado).
    Nivel senior / gerencial (ej. Coordinadora de Comunicacion en Teva).
La puntuacion es una heuristica ponderada por palabras clave del titulo (senal fuerte),
reforzada por habilidades halladas en la descripcion, con bonus de idioma/seniority
y penalizacion para roles fuera de perfil.
"""
import json, glob, os, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT_DIR = "bin/Debug/net8.0/output"

# --- solo los archivos vacantes_*.json de la sesion de hoy (20260701) ---
files = sorted(glob.glob(os.path.join(OUT_DIR, "vacantes_*_20260701_*.json")))

def norm(s):
    return (s or "").lower()

# ---------- rubrica ----------
# base por categoria de titulo (la senal mas fuerte)
TITLE_RULES = [
    (["comunicacion interna", "comunicación interna"], 93),
    (["lider de contenidos", "líder de contenidos", "content lead", "digital content"], 90),
    (["marketing de contenidos", "content marketing"], 90),
    (["content manager", "content strategist", "estratega de contenido", "gerente de contenido"], 88),
    (["especialista seo", "seo senior", "seo &", "seo/"], 85),
    (["estrategia digital", "digital strategy", "estrategia de contenido"], 80),
    (["digital marketing manager", "gerente de marketing digital", "marketing digital manager"], 78),
    (["gerente de branding", "branding & marketing", "brand manager", "gerente de marca", "coordinador de marca"], 74),
    (["gerente de marketing", "jefe de marketing", "gerente comercial / new business", "gerente de marketing y"], 72),
    (["especialista en marketing digital", "coordinador marketing digital", "coordinador de marketing", "marketing specialist"], 68),
    (["community manager", "social media", "analista de comunidad", "analista jr. de comunidad", "ejecutivo social media", "creacion de contenido y community"], 66),
    (["especialista sem", "sem "], 62),
    (["copy creativo", "guionista", "editor de contenido", "redactor", "web content editor"], 58),
    (["ejecutivo de marketing y crm", "digital marketing & crm", "digital mkt", "growth", "trafficker", "influencer marketing"], 55),
    (["analista de marketing", "analista de mercadotecnia", "analista de campanas", "analista de campañas", "consultor de marketing"], 52),
    (["marketing digital"], 60),
    (["becario de marketing", "becaria"], 30),
    (["editor de video", "productor audiovisual", "content creator", "realizador", "operadora de vizrt", "operador de", "diseñador", "disenador", "creative designer", "motion", "auxiliar de presentacion"], 25),
]

# palabras fuera de perfil -> techo de puntuacion
OFFPROFILE = ["contador", "contable", "contabilidad", "reclutador", "reclutamiento",
              "auditor", "auditoria", "auditoría", "desarrollador", ".net", "sdet",
              "devsecops", "programador", "abogado", "juridico", "jurídico",
              "almacen", "almacén", "cedis", "flotilla", "condominios",
              "comprador", "compras", "patrimonial", "vendedor sr", "vendedor senior",
              "gerente de compras", "gerente de operaciones", "gerente de auditoria",
              "gerente de cultura", "responsabilidad social", "subgerente de tienda",
              "key account", "account manager", "coordinador de ventas", "ejecutivo comercial",
              "ejecutivo de cuenta", "gerente sr de procesos", "gerente de ingenieria",
              "gerente de flotilla", "project manager senior", "senior it manager",
              "comprador senior", "capacitador de auditores", "gerente de responsabilidad",
              "desarrollo de negocios", "diseñador y editor", "productor audiovisual",
              "líder de desarrollo de negocios", "lider de desarrollo de negocios",
              "gerente de e-commerce", "gerente e commerce", "líder de crecimiento digital",
              "gerente de capacitacion"]

# habilidades del CV -> refuerzo en descripcion
SKILLS = {
    "seo": 4, "sem": 3, "keyword": 4, "posicionamiento": 3,
    "marketing de contenidos": 5, "content": 3, "contenido": 2, "editorial": 3,
    "comunicacion interna": 6, "comunicación interna": 6,
    "comunicacion corporativa": 4, "comunicación corporativa": 4,
    "email marketing": 4, "google analytics": 4, "looker": 5, "analitica": 2, "analítica": 2,
    "wordpress": 4, "joomla": 4, "photoshop": 2, "illustrator": 2, "premiere": 2, "canva": 2,
    "redaccion": 2, "redacción": 2, "copywriting": 3, "storytelling": 3,
    "kpi": 2, "metricas": 2, "métricas": 2, "branding": 2, "redes sociales": 2,
    "stakeholders": 3, "reputacion": 2, "reputación": 2,
}

LANG = ["ingles avanzado", "inglés avanzado", "bilingue", "bilingüe", "ingles fluido", "inglés fluido"]
SENIOR = ["gerente", "lider", "líder", "coordinador", "jefe", "manager", "senior", "sr"]
MKT_CTX = ["marketing", "contenido", "comunicacion", "comunicación", "seo", "social", "brand", "digital", "mercadotecnia"]

def score(titulo, desc):
    t = norm(titulo); d = norm(desc); full = t + " " + d
    # base por titulo
    base = 45
    for kws, val in TITLE_RULES:
        if any(k in t for k in kws):
            base = val
            break
    # refuerzo por habilidades (en descripcion, con tope)
    bonus = 0
    for kw, w in SKILLS.items():
        if kw in d:
            bonus += w
    bonus = min(bonus, 22)
    # idioma
    lang = 5 if any(k in full for k in LANG) else 0
    # seniority (solo si contexto de marketing/comunicacion)
    senior = 4 if (any(k in t for k in SENIOR) and any(k in full for k in MKT_CTX)) else 0
    s = base + bonus + lang + senior
    # techo si es claramente fuera de perfil (por titulo)
    if any(o in t for o in OFFPROFILE):
        s = min(s, 15)
    return max(0, min(100, round(s)))

# ---------- cargar y deduplicar ----------
term_of = {}  # (fuente,jobid) -> set de terminos de busqueda
rows = {}
for fp in files:
    base = os.path.basename(fp)
    m = re.match(r"vacantes_(occ|computrabajo)_(.+?)_ciudad-de-mexico_", base)
    term = m.group(2).replace("-", " ") if m else base
    with open(fp, encoding="utf-8") as f:
        data = json.load(f)
    for v in data:
        key = (v.get("fuente"), v.get("jobid"))
        term_of.setdefault(key, set()).add(term)
        if key not in rows:
            rows[key] = v

items = []
for key, v in rows.items():
    pts = score(v.get("titulo"), v.get("descripcion"))
    items.append((pts, v, ", ".join(sorted(term_of[key]))))
items.sort(key=lambda x: x[0], reverse=True)

# ---------- escribir excel ----------
wb = Workbook()
ws = wb.active
ws.title = "Vacantes"

headers = ["Puntuación", "Fuente", "Título", "Empresa", "Ubicación", "Salario",
           "Publicada", "Búsqueda(s)", "Enlace"]
ws.append(headers)

head_fill = PatternFill("solid", fgColor="1F4E78")
head_font = Font(bold=True, color="FFFFFF", size=11)
thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
for c in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=c)
    cell.fill = head_fill; cell.font = head_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

def color_for(p):
    if p >= 80: return "C6EFCE"   # verde
    if p >= 60: return "FFEB9C"   # amarillo
    if p >= 40: return "FCE4D6"   # naranja claro
    return "F2F2F2"               # gris

r = 2
for pts, v, terms in items:
    emp = v.get("empresa") or "Empresa confidencial"
    ws.cell(row=r, column=1, value=pts)
    ws.cell(row=r, column=2, value=v.get("fuente"))
    ws.cell(row=r, column=3, value=v.get("titulo"))
    ws.cell(row=r, column=4, value=emp)
    ws.cell(row=r, column=5, value=v.get("ubicacion"))
    ws.cell(row=r, column=6, value=v.get("salario"))
    ws.cell(row=r, column=7, value=v.get("fechaPublicacion"))
    ws.cell(row=r, column=8, value=terms)
    link = v.get("urlPublica")
    lc = ws.cell(row=r, column=9, value=link)
    lc.hyperlink = link
    lc.font = Font(color="0563C1", underline="single")
    # estilo puntuacion
    pcell = ws.cell(row=r, column=1)
    pcell.fill = PatternFill("solid", fgColor=color_for(pts))
    pcell.font = Font(bold=True)
    pcell.alignment = Alignment(horizontal="center")
    for c in range(1, len(headers) + 1):
        ws.cell(row=r, column=c).border = border
        ws.cell(row=r, column=c).alignment = Alignment(
            vertical="top",
            horizontal="center" if c in (1, 2, 7) else "left",
            wrap_text=(c in (3, 4, 5)))
    r += 1

widths = {1: 11, 2: 13, 3: 46, 4: 34, 5: 30, 6: 34, 7: 14, 8: 34, 9: 60}
for col, w in widths.items():
    ws.column_dimensions[get_column_letter(col)].width = w
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{r-1}"

# hoja de leyenda
ws2 = wb.create_sheet("Leyenda")
leg = [
    ["PUNTUACIÓN DE COINCIDENCIA — CV Claudia Elena Pedrozo Machorro", ""],
    ["", ""],
    ["Perfil evaluado", "SEO, Marketing de Contenidos, Comunicación Interna/Corporativa,"],
    ["", "Community/Social Media, Email Marketing, Google Analytics/Looker,"],
    ["", "WordPress/Joomla, Photoshop/Illustrator/Premiere/Canva."],
    ["", "Bilingüe (esp/ing avanzado). Nivel senior / gerencial."],
    ["", ""],
    ["Rango", "Interpretación"],
    ["80-100", "Coincidencia alta (rol núcleo del perfil: contenidos/SEO/comunicación)"],
    ["60-79", "Coincidencia buena (marketing digital / community / branding)"],
    ["40-59", "Coincidencia media (analista/mercadotecnia/copy/growth)"],
    ["0-39", "Baja o fuera de perfil (video, diseño, contabilidad, ventas, TI, etc.)"],
    ["", ""],
    ["Cómo se calcula", "Base por categoría del título + refuerzo por habilidades del CV"],
    ["", "halladas en la descripción + bonus idioma/seniority − penalización"],
    ["", "si el rol es claramente ajeno al perfil. Heurística orientativa."],
]
for row in leg:
    ws2.append(row)
ws2.cell(row=1, column=1).font = Font(bold=True, size=13, color="1F4E78")
ws2.cell(row=8, column=1).font = Font(bold=True)
ws2.cell(row=8, column=2).font = Font(bold=True)
ws2.column_dimensions["A"].width = 22
ws2.column_dimensions["B"].width = 78

out = os.path.join(OUT_DIR, "vacantes_puntuadas_Claudia_Pedrozo.xlsx")
wb.save(out)
print("Vacantes unicas:", len(items))
print("Guardado en:", os.path.abspath(out))
print("\nTop 12 por coincidencia:")
for pts, v, terms in items[:12]:
    print(f"  {pts:3}  [{v.get('fuente')}] {v.get('titulo')[:55]}")
