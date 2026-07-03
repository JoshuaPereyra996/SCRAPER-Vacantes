#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Ciclo de puntaje con IA (chat manual) y exportación a Excel.

Flujo:
  1. generar_prompt(sesion, cv): arma el prompt con el CV mejorado + el JSON
     consolidado de vacantes; el usuario lo pega en su IA (claude.ai, ChatGPT...).
  2. La IA devuelve un CSV: jobid,puntaje,razon (una línea por vacante).
  3. parsear_respuesta(texto): tolera encabezados, numeración y comas en la razón.
  4. exportar_excel(sesion, puntajes, ruta): Excel ordenado por puntaje con colores,
     hipervínculos, columna Estado (para el seguimiento del cliente) y leyenda.

Requiere: pip install openpyxl
"""
import json
import re

from docx_export import consolidar

# ---------------------------------------------------------------- Prompt --

PLANTILLA_PROMPT = """Eres un reclutador experto. Compara el siguiente CV con CADA vacante del JSON.

Devuelve SOLO un CSV (sin explicaciones, sin markdown, sin tabla), con este formato exacto,
una línea por vacante:

jobid,puntaje,razon

- jobid: el campo "jobid" de la vacante, tal cual.
- puntaje: entero 0-100 según el encaje del CV con la vacante (título, funciones,
  habilidades, seniority, idioma). 80-100 = encaje alto; 60-79 = bueno; 40-59 = medio;
  0-39 = bajo o fuera de perfil.
- razon: justificación de máximo 12 palabras, sin comas.

Incluye TODAS las vacantes del JSON, sin omitir ninguna.

=== CV DEL CANDIDATO ===
{cv}

=== VACANTES (JSON) ===
{vacantes}
"""


def generar_prompt(sesion: dict, cv_texto: str) -> str:
    """Construye el prompt listo para pegar en la IA."""
    vacantes = consolidar(sesion)
    # Compactar: solo los campos que la IA necesita para puntuar.
    ligeras = [{
        "jobid": v.get("jobid"),
        "fuente": v.get("fuente"),
        "titulo": v.get("titulo"),
        "empresa": v.get("empresa"),
        "ubicacion": v.get("ubicacion"),
        "salario": v.get("salario"),
        "descripcion": (v.get("descripcion") or "")[:1200],  # recortar descripciones enormes
    } for v in vacantes]
    return PLANTILLA_PROMPT.format(
        cv=cv_texto.strip() or "(CV no proporcionado: puntúa solo con el título y descripción)",
        vacantes=json.dumps(ligeras, ensure_ascii=False, indent=1))


# ------------------------------------------------------- Parseo respuesta --

def parsear_respuesta(texto: str) -> dict:
    """
    Convierte la respuesta CSV de la IA en {jobid: (puntaje, razon)}.
    Tolerante: ignora encabezados, viñetas, numeración, ``` de markdown y
    acepta ',' o ';' como separador. La razón puede contener comas.
    """
    puntajes = {}
    for linea in texto.splitlines():
        linea = linea.strip().strip("`").lstrip("-*• ").strip()
        if not linea or linea.lower().startswith(("jobid", "csv", "```")):
            continue
        sep = ";" if (";" in linea and "," not in linea.split(";")[0]) else ","
        partes = linea.split(sep, 2)
        if len(partes) < 2:
            continue
        jobid = partes[0].strip().strip('"').strip()
        m = re.search(r"\d{1,3}", partes[1])
        if not jobid or not m:
            continue
        puntaje = max(0, min(100, int(m.group())))
        razon = partes[2].strip().strip('"').strip() if len(partes) > 2 else ""
        puntajes[jobid] = (puntaje, razon)
    return puntajes


# ------------------------------------------------------------------ Excel --

def exportar_excel(sesion: dict, puntajes: dict, ruta_xlsx: str) -> tuple:
    """
    Genera el Excel puntuado. Devuelve (total_vacantes, con_puntaje).
    Las vacantes sin puntaje de la IA aparecen al final con puntaje vacío.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    vacantes = consolidar(sesion)
    filas = []
    for v in vacantes:
        p = puntajes.get(str(v.get("jobid")))
        filas.append((p[0] if p else None, p[1] if p else "", v))
    # Ordenar: puntuadas de mayor a menor, sin puntaje al final.
    filas.sort(key=lambda x: (x[0] is None, -(x[0] or 0)))

    wb = Workbook()
    ws = wb.active
    ws.title = "Vacantes"

    encabezados = ["Puntaje", "Fuente", "Título", "Empresa", "Ubicación", "Salario",
                   "Publicada", "Razón (IA)", "Estado", "Enlace"]
    ws.append(encabezados)

    relleno_enc = PatternFill("solid", fgColor="1F4E78")
    fuente_enc = Font(bold=True, color="FFFFFF", size=11)
    fino = Side(style="thin", color="D9D9D9")
    borde = Border(left=fino, right=fino, top=fino, bottom=fino)
    for c in range(1, len(encabezados) + 1):
        celda = ws.cell(row=1, column=c)
        celda.fill = relleno_enc
        celda.font = fuente_enc
        celda.alignment = Alignment(horizontal="center", vertical="center")
        celda.border = borde

    def color(p):
        if p is None: return "F2F2F2"
        if p >= 80: return "C6EFCE"    # verde
        if p >= 60: return "FFEB9C"    # amarillo
        if p >= 40: return "FCE4D6"    # naranja claro
        return "F2F2F2"                # gris

    # Lista desplegable para la columna Estado (seguimiento del cliente).
    dv = DataValidation(
        type="list",
        formula1='"Por revisar,Aplicada,Descartada,Entrevista,Oferta"',
        allow_blank=True)
    ws.add_data_validation(dv)

    r = 2
    for puntaje, razon, v in filas:
        ws.cell(row=r, column=1, value=puntaje)
        ws.cell(row=r, column=2, value=v.get("fuente"))
        ws.cell(row=r, column=3, value=v.get("titulo"))
        ws.cell(row=r, column=4, value=v.get("empresa") or "Empresa confidencial")
        ws.cell(row=r, column=5, value=v.get("ubicacion"))
        ws.cell(row=r, column=6, value=v.get("salario"))
        ws.cell(row=r, column=7, value=v.get("fechaPublicacion"))
        ws.cell(row=r, column=8, value=razon)
        estado = ws.cell(row=r, column=9, value="Por revisar")
        dv.add(estado)
        enlace = v.get("urlPublica")
        celda_enlace = ws.cell(row=r, column=10, value=enlace)
        if enlace:
            celda_enlace.hyperlink = enlace
            celda_enlace.font = Font(color="0563C1", underline="single")

        celda_p = ws.cell(row=r, column=1)
        celda_p.fill = PatternFill("solid", fgColor=color(puntaje))
        celda_p.font = Font(bold=True)
        celda_p.alignment = Alignment(horizontal="center")
        for c in range(1, len(encabezados) + 1):
            ws.cell(row=r, column=c).border = borde
            ws.cell(row=r, column=c).alignment = Alignment(
                vertical="top",
                horizontal="center" if c in (1, 2, 7, 9) else "left",
                wrap_text=(c in (3, 4, 5, 8)))
        r += 1

    anchos = {1: 10, 2: 13, 3: 44, 4: 30, 5: 26, 6: 26, 7: 13, 8: 38, 9: 13, 10: 55}
    for col, w in anchos.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(encabezados))}{r - 1}"

    # Hoja de leyenda.
    ws2 = wb.create_sheet("Leyenda")
    leyenda = [
        ["PUNTAJE DE COINCIDENCIA (calculado por IA contra el CV del candidato)", ""],
        ["", ""],
        ["Rango", "Interpretación"],
        ["80-100", "Coincidencia alta: aplicar con prioridad"],
        ["60-79", "Coincidencia buena"],
        ["40-59", "Coincidencia media"],
        ["0-39", "Baja o fuera de perfil"],
        ["(vacío)", "La IA no devolvió puntaje para esa vacante"],
        ["", ""],
        ["Columna Estado", "Úsala para llevar tu seguimiento: Por revisar / Aplicada /"],
        ["", "Descartada / Entrevista / Oferta (lista desplegable)."],
    ]
    for fila in leyenda:
        ws2.append(fila)
    ws2.cell(row=1, column=1).font = Font(bold=True, size=13, color="1F4E78")
    ws2.cell(row=3, column=1).font = Font(bold=True)
    ws2.cell(row=3, column=2).font = Font(bold=True)
    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 70

    wb.save(ruta_xlsx)
    con_puntaje = sum(1 for p, _, _ in filas if p is not None)
    return len(filas), con_puntaje
